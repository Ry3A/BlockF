import vk_api
import re
import openpyxl
import requests
import datetime
import time
import sqlite3

import PIL.Image as Image
from bs4 import BeautifulSoup
from matplotlib import pyplot as plt
from vk_api import VkUpload
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id

MIREA_URL: str = 'https://www.mirea.ru/schedule/'


def show_hi_message(user_id):
    """
    Выводит приветственное сообщение, предлагает выбрать группу

    """
    text = f"Привет, {vk.users.get(user_id=user_id)[0]['first_name']}!\nНапиши номер своей группы!"
    send_message(user_id, text)


def get_string_date(date, with_week_day=False):
    """
    Преобразует дату в строку с датой

    """
    result = ''
    WEEK_DAYS = ['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу', 'воскресенье']
    MONTHS_DAYS = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября',
                   'ноября', 'декабря']
    if with_week_day:
        result += WEEK_DAYS[date.isocalendar().weekday - 1] + " "
    result += str(date.day) + " " + MONTHS_DAYS[date.month % 12 - 1]
    return result


def show_schedule_period_keyboard(user_id):
    keyboard = VkKeyboard(one_time=False)
    keyboard.add_button("на сегодня", color=VkKeyboardColor.POSITIVE)
    keyboard.add_button("на завтра", color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button("на эту неделю", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("на следующую неделю", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("какая неделя?", color=VkKeyboardColor.SECONDARY)
    keyboard.add_button("какая группа?", color=VkKeyboardColor.SECONDARY)
    keyboard.add_button("Помощь", color=VkKeyboardColor.SECONDARY)
    send_message_keyboard(user_id, 'Показать расписание ...', keyboard)


def show_help(user_id):
    send_message(user_id,
                 "Список команд:\n\nНачать - Запускает бота\nБот - Выбор периода для показа расписания\nКорона - Выбор периода для показа расписания\n")


def show_save_group(user_id, group):
    edit_user_group(user_id, group)
    print('Сохраняем группу')
    # cur.execute("SELECT * FROM groups WHERE user_id = '{}'".format(user_id))


def get_group_column(group):
    """
    Ищет столбец группы в расписании

    """
    for i in range(0, len(schedule_data), 4):
        if schedule_data[i][0] == group:
            return i
    return -1


def search_group(user_id, group):
    group = group.upper()
    pattern = r'\w{4}-\d{2}-\d{2}'
    if re.match(pattern, group):
        if get_group_column(group) != -1:
            return True
    return False


def show_corona_ru_stat(user_id):
    page = requests.get('https://coronavirusstat.ru/country/russia')  # Получаем страницу
    soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    result = soup.find('table', {'class': 'table table-bordered small'}).findAll('tr')
    days = []
    active = []
    cured = []
    died = []
    cases = []
    stats = []
    ml = 1000000

    for i in range(1, 11):
        days.append(result[i].find('th').getText())
        for a in result[i].findAll('td'):
            stats.append(int(a.getText().split(' ')[1]))
    for i in range(0, len(stats), 4):
        active.append(stats[i] / ml)
    for i in range(1, len(stats), 4):
        cured.append(stats[i] / ml)
    for i in range(2, len(stats), 4):
        died.append(stats[i] / ml)
    for i in range(3, len(stats), 4):
        cases.append(stats[i] / ml)

    days = list(reversed(days))
    active = list(reversed(active))
    cured = list(reversed(cured))
    died = list(reversed(died))
    cases = list(reversed(cases))
    graf_data = {
        'Активных': active,
        'Вылечено': cured,
        'Умерло': died,
    }
    for i in range(len(days)):
        days[i] = days[i][:-5]
    fig, ax = plt.subplots()
    ax.stackplot(days, graf_data.values(),
                 labels=graf_data.keys(), alpha=0.8)
    ax.legend(loc='upper left')
    ax.set_title("123")
    ax.set_ylabel("456")
    fig.savefig('graf.png')
    upload = VkUpload(vk_session)
    attachments = []
    photo = upload.photo_messages("graf.png")[0]
    attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
    send_message_attachments(user_id=user_id,
                             text=reformat_corona_data('Россия', get_corona_stat()),
                             attachments=attachments)


def get_corona_stat(extra_url=''):
    """
    Возвращает статистику коронавируса на сегодня определённой области
    :param extra_url:
    :return:
    """
    page = requests.get('https://coronavirusstat.ru' + extra_url)  # Получаем страницу
    soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    result = soup.find(string='Прогноз заражения на 10 дней').find_parent('div', {
        'class': 'border rounded mt-3 mb-3 p-3'})
    status = result.find('h6', 'text-muted').getText()[:-17]
    data = result.findAll('div', {'class': 'col col-6 col-md-3 pt-4'})
    plus = [] * 4
    value = [] * 4
    for i in range(4):
        value.append(data[i].find('div', 'h2').getText())
        plus.append(data[i].find('span', {'class': 'font-weight-bold'}).getText())
    return status, value, plus


def show_corona_region_stat(user_id, rg):
    if len(rg) > 0:
        region = rg[0].title()
        page = requests.get('https://coronavirusstat.ru')  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.findAll('div', {'class': 'c_search_row'})
        d = ''
        rg = 'Не найден'
        for x in result:
            tmp = x.find('span', 'small').find('a')
            if region.title() in tmp.getText().split(' '):
                rg = tmp.getText()
                d = tmp.get('href')
                break
        send_message(user_id, reformat_corona_data(rg, get_corona_stat(d)))


def reformat_corona_data(region, data):
    status, value, plus = data
    return '{}\n\nРегион: {}\nСлучаев: {} ({} за сегодня)\nАктивных: {} ({} за сегодня)\n' \
           'Вылечено: {} ({} за сегодня)\nУмерло: {} ({} за сегодня)'.format(status, region, value[0], plus[0],
                                                                             value[1], plus[1],
                                                                             value[2], plus[2], value[3], plus[3])


def send_message_attachments(user_id, text, attachments=list):
    vk.messages.send(
        user_id=user_id,
        attachment=','.join(attachments),
        random_id=get_random_id(),
        message=text
    )


def show_weather_keyboard(user_id):
    keyboard1 = VkKeyboard(one_time=False)
    keyboard1.add_button("сейчас", color=VkKeyboardColor.PRIMARY)
    keyboard1.add_button("сегодня", color=VkKeyboardColor.POSITIVE)
    keyboard1.add_button("завтра", color=VkKeyboardColor.POSITIVE)
    keyboard1.add_line()
    keyboard1.add_button("на 5 дней", color=VkKeyboardColor.POSITIVE)
    send_message_keyboard(user_id, 'Показать погоду в Москве ...', keyboard1)


def getPic(n):
    token = "6405692b910d4f8e7819f42ee8b12caa"
    r = requests.get(f"http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid={token}&units=metric&lang=ru")
    data = r.json()
    if n == 1:
        a = 0
        b = 4
        c = 1
    if n == 2:
        a = 4
        b = 8
        c = 1
    if n == 3:
        a = 1
        b = 20
        c = 4
    for i in range(a, b, c):
        icon = data["list"][i]["weather"][0]["icon"]
        image = requests.get(f"http://openweathermap.org/img/wn/{icon}@2x.png", stream=True)
        with open(f"file{i}.png", "wb") as f:
            f.write(image.content)
    if n == 1 or n == 2:
        img = Image.new('RGBA', (400, 100))
    elif n == 3:
        img = Image.new('RGBA', (500, 100))
    img0 = Image.open("file0.png")
    img1 = Image.open("file1.png")
    img2 = Image.open("file2.png")
    img3 = Image.open("file3.png")
    img.paste(img0, (0, 0))
    img.paste(img1, (100, 0))
    img.paste(img2, (200, 0))
    img.paste(img3, (300, 0))
    if n == 3:
        img4 = Image.open("file4.png")
        img.paste(img4, (400, 0))
    img.save("image.png")


def show_today_weather(user_id):
    getPic(1)
    attachments = []
    photo = upload.photo_messages("image.png")[0]
    attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
    vk.messages.send(
        user_id=event.user_id,
        attachment=','.join(attachments),
        random_id=get_random_id(),
        message="Погода в Москве сегодня\n")
    send_message(user_id, getWeater_day(1))


def show_tomorrow_weather(user_id):
    getPic(2)
    attachments = []
    photo = upload.photo_messages("image.png")[0]
    attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
    vk.messages.send(
        user_id=event.user_id,
        attachment=','.join(attachments),
        random_id=get_random_id(),
        message="Погода в Москве завтра\n")
    send_message(user_id, getWeater_day(2))


def getWeater_day5():
    token = "6405692b910d4f8e7819f42ee8b12caa"
    r = requests.get(f"http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid={token}&units=metric&lang=ru")
    data = r.json()
    temp_1 = ""
    temp_2 = ""
    for i in range(1, 20, 4):
        temp_1 += "/" + str(data["list"][i]["main"]["temp"]) + "/"
    for i in range(2, 20, 4):
        temp_2 += "/" + str(data["list"][i]["main"]["temp"]) + "/"
    temp_1 += "ДЕНЬ\n"
    temp_2 += "НОЧЬ"
    return (temp_1 + temp_2)


def show_five_days_weather(user_id):
    day_1 = datetime.datetime.now().date()
    day_2 = day_1 + datetime.timedelta(days=4)
    day_1 = day_1.strftime("%d.%m")
    day_2 = day_2.strftime("%d.%m")
    getPic(3)
    attachments = []
    photo = upload.photo_messages("image.png")[0]
    attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

    vk.messages.send(
        user_id=event.user_id,
        attachment=','.join(attachments),
        random_id=get_random_id(),
        message=f"Погода в Москве c {day_1} по {day_2}\n")
    send_message(user_id, getWeater_day5())


def getWeater_day(n):
    token = "6405692b910d4f8e7819f42ee8b12caa"
    r = requests.get(f"http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid={token}&units=metric&lang=ru")
    data = r.json()
    temp = ""
    info = ""
    day = ["УТРО", "ДЕНЬ", "ВЕЧЕР", "НОЧЬ"]
    if n == 1:
        a = 0
        b = 4
    if n == 2:
        a = 4
        b = 8
    for i in range(a, b):
        temp += "/" + str(data["list"][i]["main"]["temp"]) + "/"
        info += f'{day[i - a]}\n'
        info += f'//{(data["list"][i]["weather"][0]["description"])}, температура: {str(data["list"][i]["main"]["temp_min"])} - {str(data["list"][i]["main"]["temp_max"])}°С\n'
        info += f'//Давление: {str(data["list"][i]["main"]["pressure"])} мм рт. ст., влажность{str(data["list"][i]["main"]["humidity"])}%\n'
        info += f'//Ветер: {(data["list"][i]["wind"]["speed"])}, {str(data["list"][i]["wind"]["speed"])} м/с, {(data["list"][i]["wind"]["deg"])}\n'
    temp += "\n\n"
    return (temp + info)


def show_weather_today(user_id):
    r = requests.get(
        f'http://api.openweathermap.org/data/2.5/weather?q=moscow&appid=6405692b910d4f8e7819f42ee8b12caa&lang=ru'
        f'&units=metric')
    data = r.json()
    stepen = ''
    direction = ''
    weather = data['weather'][0]['description']
    temprature_min = data['main']['temp_min']
    temprature_max = data['main']['temp_max']
    pressure = data['main']['pressure']
    humidity = data['main']['humidity']
    wind = data['wind']['speed']
    degree = data['wind']['deg']
    if wind < 0.2:
        stepen = 'штиль'
    elif 0.3 < wind < 1.6:
        stepen = 'тихий'
    elif 1.5 < wind < 3.4:
        stepen = 'легкий'
    elif 3.3 < wind < 5.5:
        stepen = 'слабый'
    elif 5.4 < wind < 8:
        stepen = 'умеренный'
    elif 7.9 < wind < 10.8:
        stepen = 'свежий'
    elif 13.8 < wind < 17.2:
        stepen = 'крепкий'
    elif 17.1 < wind < 20.8:
        stepen = 'очень крепкий'
    elif 20.7 < wind < 24.5:
        stepen = 'шторм'
    elif 24.4 < wind < 28.5:
        stepen = 'сильный шторм'
    elif 28.4 < wind < 32.6:
        stepen = 'жестокий шторм'
    elif wind > 32.9:
        stepen = 'ураган'
    if (degree < 22.51) or ((degree - 315 > 360 - degree) and (315 < degree < 360)):
        direction = 'северный'
    elif (22.5 < degree < 45) or ((degree - 45 < 90 - degree) and (45 < degree < 90)):
        direction = 'северо-восточный'
    elif ((degree - 45 > 90 - degree) and (45 < degree < 90)) or ((degree - 90 < 135 - degree) and (90 < degree < 135)):
        direction = 'восточный'
    elif ((degree - 90 > 135 - degree) and (90 < degree < 135)) or (
            (degree - 135 < 180 - degree) and (135 < degree < 180)):
        direction = 'юго-восточный'
    elif ((degree - 135 > 180 - degree) and (135 < degree < 180)) or (
            (degree - 180 < 225 - degree) and (180 < degree < 225)):
        direction = 'южный'
    elif ((degree - 180 > 225 - degree) and (180 < degree < 225)) or (
            (degree - 225 < 270 - degree) and (225 < degree < 270)):
        direction = 'юго-западный'
    elif ((degree - 225 > 270 - degree) and (225 < degree < 270)) or (
            (degree - 270 < 315 - degree) and (270 < degree < 315)):
        direction = 'западный'
    elif ((degree - 270 > 315 - degree) and (270 < degree < 315)) or (
            (degree - 315 < 360 - degree) and (315 < degree < 360)):
        direction = 'северо-западный'
    send_message(user_id, f'{weather}\nТемпература: {round(temprature_min)}...{round(temprature_max)}°C\n'
                          f'Давление: {pressure} мм.рт.ст., влажность: {humidity}%\n'
                          f'Ветер: {stepen}, {wind} м/с, {direction}\n')

def show_current_week(user_id):
    send_message(user_id, "Текущая неделя - {}".format(get_current_week()))


def do_command(user_id, text):
    for i in range(len(users_to_get_teacher)):
        if users_to_get_teacher[i][0] == user_id:
            if text == 'на сегодня':
                show_today_teacher_schedule(user_id, users_to_get_teacher[i][1])
            elif text == 'на завтра':
                show_today_teacher_schedule(user_id, users_to_get_teacher[i][1], 1)
            elif text == 'на эту неделю':
                show_teacher_week_schedule(user_id, users_to_get_teacher[i][1])
            elif text == 'на следующую неделю':
                show_teacher_week_schedule(user_id, users_to_get_teacher[i][1], 1)
            del users_to_get_teacher[i]
            return
    cmds = {
        'начать': show_hi_message,
        'бот': show_schedule_period_keyboard,
        'помощь': show_help,
        'корона': show_corona_ru_stat,
        'погода': show_weather_keyboard,
        'сейчас': show_weather_today,
        'сегодня': show_today_weather,
        'завтра': show_tomorrow_weather,
        'на 5 дней': show_five_days_weather,
        'на сегодня': show_today_schedule,
        'на завтра': show_tomorrow_schedule,
        'на эту неделю': show_week_schedule,
        'какая неделя?': show_current_week,
        'какая группа?': show_user_group
    }
    if text in cmds:
        cmds[text](user_id)
        return
    if text == 'на следующую неделю':
        show_week_schedule(user_id, 1)
        return

    msg = text.split(' ')
    cmds2 = {
        'корона': show_corona_region_stat,
        'найти': show_teacher_keyboard
    }
    if msg[0] in cmds2:
        cmds2[msg[0]](user_id, msg[1:])
        return
    # ...

    if search_group(user_id, text):
        show_save_group(user_id, group=text)
        return
    send_message(user_id, "Команда не найдена :/ \nДля списка команд напишите: Помощь")

def show_user_group(user_id):
    group = get_user_group(user_id)
    if group:
        send_message(user_id, 'Твоя группа - {}'.format(group))


def send_message(user_id, text):
    vk.messages.send(
        user_id=user_id,
        random_id=get_random_id(),
        message=text
    )


def send_message_keyboard(user_id, text, keyboard):
    vk.messages.send(
        user_id=user_id,
        random_id=get_random_id(),
        keyboard=keyboard.get_keyboard(),
        message=text
    )


def get_week_schedule(group, date, with_reformat=True):
    """
    Возвращает расписанию на неделю, дату которой передали

    """
    now = date.isocalendar()
    week = now.week - 5
    week_even = (week + 1) % 2  # Является ли неделя чётной
    column = get_group_column(group)
    out = []
    tmp = []
    for i in range(2 + week_even, len(schedule_data[column]), 2):  # Каждый второй, со смещением по недели
        tmp.append(
            [schedule_data[column][i],  # Предмет
             schedule_data[column + 1][i],  # Вид занятий
             schedule_data[column + 2][i],  # Преподаватель
             schedule_data[column + 3][i]]  # Кабинет
        )
        if len(tmp) == 6:
            out.append(tmp)
            tmp = []

    for i in range(len(out)):
        for j in range(6):
            out[i][j][0] = reformat_subject_name(out[i][j][0], week_number=week,
                                                 ignore_weeks=(not with_reformat))
            out[i][j][1] = reformat_double_pair(out[i][j][1])
            out[i][j][2] = reformat_double_pair(out[i][j][2])
            out[i][j][3] = reformat_double_pair(out[i][j][3])
    return out


def get_day_schedule(group, date):
    """
    Возвращает массив с расписанием на текущий день

    """
    week = get_week_schedule(group, date)
    week_index = date.isocalendar().weekday - 1
    if week_index == 6:
        return [[] * 4] * 6
    return week[week_index]


def get_current_week():
    """
    Возвращает номер текущей недели

    """
    return datetime.datetime.now().isocalendar().week - 5


def reformat_subject_name(name, week_number, ignore_weeks=False):
    """
    Реформат названия предмета с проверкой его присутствия на определённой неделе

    """
    custom_week_pattern = r'кр. ([\d\,]+) н. ([^\\]+)'  # Кроме каких-то недель
    custom_week_range_pattern = r'(\d+\-\d+) н. ([^\\]+)'  # Диапазон
    custom_week_is_set_pattern = r'([\d\,]+) н. ([^\\]+)'  # Включая эти недели
    custom_week_dirt_pattern = r'…'  # Заглушки в расписании
    if name and name != 'None':  # Пара есть?
        data = name.split('\n')
        # Цикл, для сдвоенных пар
        for i in range(len(data)):
            if not ignore_weeks:
                kr = re.search(custom_week_pattern, data[i])  # Проверяем, есть ли паттерн КР
                if kr:
                    if str(week_number) in kr.group(1).split(','):  # Если неделя в списке исключённых удаляем
                        data[i] = '--'
                    else:
                        data[i] = kr.group(2)
                else:
                    range_week = re.search(custom_week_range_pattern, data[i])
                    if range_week:
                        tmp = range_week.group(1).split('-')
                        from_week = int(tmp[0])
                        to_week = int(tmp[1])
                        if from_week <= week_number <= to_week:
                            data[i] = range_week.group(2)
                        else:
                            data[i] = '--'
                    else:
                        is_set = re.search(custom_week_is_set_pattern, data[i])
                        if is_set:
                            if str(week_number) in is_set.group(1).split(','):
                                data[i] = is_set.group(2)
                            else:
                                data[i] = '--'
                        else:
                            dirt = re.search(custom_week_dirt_pattern, data[i])
                            if dirt:
                                data[i] = '--'
        return ' / '.join(data) if data else '--'
    return '--'


def reformat_double_pair(data):
    """
    Двойные и пустые пары в читабельный формат

    """
    if data:
        if data == 'None':
            return ''
        return ' / '.join(data.split('\n'))
    return '--'


def get_teacher_full_name(teacher):
    """
    Получение полного имени преподавателей из расписания

    """
    result = set()
    for i in range(2, len(schedule_data), 4):
        for j in range(2, len(schedule_data[i])):
            tmp = schedule_data[i][j].split('\n')
            if len(tmp) > 0:
                if tmp[0].split(' ')[0] == teacher:
                    result.add(tmp[0] if tmp[0][-1] == '.' else tmp[0] + '.')  # Исправление косяков расписания
                elif tmp[-1].split(' ')[0] == teacher:

                    result.add(tmp[-1] if tmp[-1][-1] == '.' else tmp[-1] + '.')
    return result


def get_teacher_week_schedule(teacher, date, with_reformat=True):
    """
    Возвращает расписание преподавателя на указанную неделю

    """
    now = date.isocalendar()
    week = now.week - 5
    week_even = (week + 1) % 2  # Является ли неделя чётной
    out = []
    tmp = []

    for j in range(2 + week_even, len(schedule_data[0]), 2):
        para = []  # одна пара
        for i in range(2, len(schedule_data), 4):  # Слева на права
            tmp_teachers = schedule_data[i][j].split('\n')  # для сдвоенных пар
            if len(tmp_teachers) > 0:
                t1 = tmp_teachers[0] if tmp_teachers[0][-1] == '.' else tmp_teachers[0] + '.'
                t2 = tmp_teachers[-1] if tmp_teachers[-1][-1] == '.' else tmp_teachers[-1] + '.'
                if t1 == teacher:
                    para = [
                        schedule_data[i - 2][j].split('\n')[0],  # Предмет
                        schedule_data[i - 1][j].split('\n')[0],  # Вид
                        schedule_data[i - 2][0],  # Группа
                        schedule_data[i + 1][j].split('\n')[0]  # Аудитория
                    ]
                    break
                elif t2 == teacher:
                    para = [
                        schedule_data[i - 2][j].split('\n')[-1],  # Предмет
                        schedule_data[i - 1][j].split('\n')[-1],  # Вид
                        schedule_data[i - 2][0],  # Группа
                        schedule_data[i + 1][j].split('\n')[-1]  # Аудитория
                    ]
                    break
                # Останавливаем смешение вправо, если нашли
        tmp.append(para)  # Добавляем пару, даже если она пустая
        if (j - week_even) % 12 == 0:
            out.append(tmp)
            tmp = []

    for i in range(len(out)):
        for j in range(6):
            if len(out[i][j]) > 1:
                out[i][j][0] = reformat_subject_name(out[i][j][0], week_number=week,
                                                     ignore_weeks=(not with_reformat))
    return out


def get_day_teacher_schedule(teacher, date):
    """
    Возвращает расписание преподавателя на переданный день

    """
    week = get_teacher_week_schedule(teacher, date)
    week_index = date.isocalendar().weekday - 1
    if week_index == 6:
        return [[] * 4] * 6
    return week[week_index]


def reformat_day_schedule(data, date: datetime.datetime = datetime.datetime.now(),
                          week_format: bool = False, with_header: bool = True,
                          teacher_header=None):
    """
    Форматирует один день из списка в строку для дальнейшего вывода

    """
    result = ''
    if with_header:
        if teacher_header:
            result += '\nРасписание преподавателя {} на {}:\n'.format(
                teacher_header, get_string_date(date, with_week_day=week_format))  # Дата
        else:
            result += '\nРасписание на {}:\n'.format(
                get_string_date(date, with_week_day=week_format))  # Дата
    for i in range(len(data)):
        if len(data[i]) > 1:
            if data[i][0][:len('--')] != '--':
                result += '{}) {}, {}, {}, {}\n'.format(
                    i + 1,
                    str(data[i][0]),
                    str(data[i][1]) if data[i][1] != '--' and \
                                       data[i][1] != '' else '_',
                    str(data[i][2]) if data[i][2] != '--' and \
                                       data[i][2] != '' else '_',
                    str(data[i][3])) if data[i][3] != '--' and \
                                        data[i][3] != '' else '_'
            else:
                result += '{}) {}\n'.format(i + 1, '--')
        else:
            result += '{}) {}\n'.format(i + 1, '--')
    return result


def reformat_teacher_name(teacher):
    """
    Форматирует имя преподавателя

    """
    tmp = teacher.split(' ')
    if len(tmp) == 2:
        teacher = tmp[0].title() + ' ' + tmp[1].upper()
    elif len(tmp) == 1:
        teacher = tmp[0].title()
    return teacher


def show_today_teacher_schedule(user_id, teacher, day_delta=0):
    """
    Выводит расписания на сегодня день(со смещением)

    """
    now = datetime.datetime.now() + datetime.timedelta(days=day_delta)
    teacher = reformat_teacher_name(teacher)
    if validate_teacher_name(teacher):
        schedule = get_day_teacher_schedule(teacher, now)
        send_message(user_id=user_id, text=reformat_day_schedule(schedule, now, teacher_header=teacher))


def show_teacher_week_schedule(user_id, teacher, week_delta=0):
    """
    Выводит расписание преподавателя на неделю

    """
    now = datetime.datetime.now() + datetime.timedelta(weeks=week_delta)
    day_date = now - datetime.timedelta(days=now.isocalendar().weekday - 1)
    result = ''
    teacher = reformat_teacher_name(teacher)
    if validate_teacher_name(teacher):
        schedule = get_teacher_week_schedule(teacher, now)
        for i in range(6):
            result += reformat_day_schedule(schedule[i], date=day_date, teacher_header=teacher,
                                            week_format=True)
            day_date += datetime.timedelta(days=1)
        send_message(user_id, result)


def show_teacher_period_keyboard(user_id, teacher):
    """
    Показывает клавиатуру для выбора периода расписания преподавателя

    """
    tmp = teacher.split(' ')
    if len(tmp) == 2:
        teacher = tmp[0].title() + ' ' + tmp[1].upper()
        if validate_teacher_name(teacher):
            # Создаём клавиатуру
            keyboard = VkKeyboard(one_time=True)
            keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
            keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
            keyboard.add_line()
            keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
            keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)

            clear_wait_lists(user_id)
            add_user_to_get_teacher_list(user_id, teacher)
            send_message_keyboard(user_id, text='Показать расписание преподавателя {} ...'.format(teacher),
                                  keyboard=keyboard)
            return
    send_message(user_id, text='Преподаватель не найден')


def _show_week_day_schedule(user_id, day):
    """
    Показывает расписание на определённый день недели

    """
    WEEK_INFINITIVE = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']
    group = get_user_group(user_id)
    if group:
        date = datetime.datetime.now()
        if get_current_week() % 2 == 0:  # Если она не чётная
            date -= datetime.timedelta(weeks=1)
        odd = get_week_schedule(group=group, date=date, with_reformat=False)
        date += datetime.timedelta(weeks=1)
        even = get_week_schedule(group=group, date=date, with_reformat=False)
        index = WEEK_INFINITIVE.index(day)
        o = reformat_day_schedule(data=odd[index], with_header=False)  # Нечётный день
        e = reformat_day_schedule(data=even[index], with_header=False)  # Чётный день
        result = 'Расписание на {}, нечётной недели\n'.format(
            day.title()) + o + '\n\n' + 'Расписание на {} чётной недели\n'.format(day.title()) + e
        send_message(user_id, result)


def get_user_group(user_id):
    """
    Получает группу пользователя или ошибка

    :param user_id:
    :return: Номер группы или None
    """
    for i in range(len(users_groups)):
        if users_groups[i][0] == user_id:
            return users_groups[i][1]
    send_message(user_id, "Не сохранена группа!")
    return None


def show_today_schedule(user_id):
    """
    Выводит расписание на сегодняшний день

    """
    now = datetime.datetime.now()
    group = get_user_group(user_id)
    if group:
        schedule = get_day_schedule(group, now)
        send_message(user_id=user_id, text=reformat_day_schedule(schedule, now))


def show_tomorrow_schedule(user_id):
    """
    Выводит расписание на завтрашний день

    """
    now = datetime.datetime.now() + datetime.timedelta(days=1)
    group = get_user_group(user_id)
    if group:
        schedule = get_day_schedule(group, now)
        send_message(user_id=user_id, text=reformat_day_schedule(schedule, now))


def show_week_schedule(user_id, week_delta=0):
    """
    Выводит расписание на текущую неделю

    """
    now = datetime.datetime.now() + datetime.timedelta(weeks=week_delta)
    day_date = now - datetime.timedelta(days=now.isocalendar().weekday - 1)
    group = get_user_group(user_id)
    result = ''
    if group:
        schedule = get_week_schedule(group, now)
        for i in range(6):
            result += reformat_day_schedule(schedule[i], date=day_date, week_format=True)
            day_date += datetime.timedelta(days=1)
        send_message(user_id, result)


def show_teacher_keyboard(user_id, teacher=None):
    """
    Показывает клавиатуру выбора преподавателя, либо выбора периода, для показа расписания

    """
    name = ''
    if len(teacher) == 2:
        name = teacher[0].title() + ' ' + teacher[1].upper()
    elif len(teacher) == 1:  # Только фамилия
        tmp = []
        for a in get_teacher_full_name(teacher[0].title()):
            tmp.append(a)
        if len(tmp) == 1:  # Если 1, то
            name = tmp[0]
        elif len(tmp) > 1:
            add_user_to_set_teacher_list(user_id)  # Добавляем пользователя в список ожидания
            keyboard = VkKeyboard(one_time=True)
            for i in range(len(tmp)):
                keyboard.add_button(tmp[i], color=VkKeyboardColor.SECONDARY)
                if i % 2 and i != len(tmp) - 1:  # Каждый второй, но не последний
                    keyboard.add_line()
            send_message_keyboard(user_id=user_id, text='Выберите преподавателя', keyboard=keyboard)
            return
    if len(name) > 1:
        show_teacher_period_keyboard(user_id, name)
        return
    send_message(user_id, 'Преподаватель не найден')


def add_user_to_edit_group_list(user_id):
    """
    Добавляет пользователя в список обновления группы

    """
    users_to_set_group.add(str(user_id))


def add_user_to_set_teacher_list(user_id):
    """
    Добавляет пользователя в список выбора преподавателя

    """
    users_to_set_teacher.add(str(user_id))




def add_user_to_get_teacher_list(user_id, teacher):
    """
    Добавляет пользователя в список выбора преподавателя

    """
    users_to_get_teacher.append([user_id, teacher])


def edit_user_group(user_id, group_slug):
    """
    Изменить группу пользователя или выдать ошибку

    """
    group_slug = group_slug.upper()
    if validate_group_slug(group_slug):
        flag = False
        for i in range(len(users_groups)):
            if users_groups[i][0] == user_id:
                flag = True
                users_groups[i][1] = group_slug
                break
        if not flag:
            users_groups.append([user_id, group_slug])
        send_message(user_id, 'Я запомнил, что ты учишься в группе {}'.format(group_slug))
        show_schedule_period_keyboard(user_id=user_id)  # Показываем клавиатуры выбора
    else:
        send_message(user_id, 'Неверный формат или группа не найдена!\n\nФормат: \'АБВГ-12-34\'')


def clear_wait_lists(user_id):
    """
    Убирает пользователя из списка ожидания

    :param user_id:
    :return:
    """
    users_to_set_group.discard(str(user_id))
    users_to_set_teacher.discard(str(user_id))
    for i in range(len(users_to_get_teacher)):
        if users_to_get_teacher[i][0] == user_id:
            del users_to_get_teacher[i]
            break


def validate_teacher_name(teacher):
    """
    Проверяет наличие преподавателя в файлах расписания

    """
    for i in range(2, len(schedule_data), 4):
        for j in range(2, len(schedule_data[i])):
            tmp = schedule_data[i][j].split('\n')  # для сдвоенных пар
            if len(tmp) > 0:
                # Исправляем отсутствие точки у некоторых преподавателей
                t1 = tmp[0] if tmp[0][-1] == '.' else tmp[0] + '.'
                t2 = tmp[-1] if tmp[-1][-1] == '.' else tmp[-1] + '.'
                if t1 == teacher or t2 == teacher:
                    return True
    return False


def validate_group_slug(group_slug):
    """
    Проверка на валидность номера группы по маске и списку групп

    """
    group_slug = group_slug.upper()
    patterngroop = r'\w{4}-\d{2}-\d{2}'
    if re.match(patterngroop, group_slug):
        if get_group_column(group_slug) != -1:
            return True
    return False


def schedule_file():
    """
    Парсит полученные файлы расписание и записывает в списки
    """

    pattern = r'\w{4}-\d{2}-\d{2}'
    for c in range(3):
        book = openpyxl.load_workbook(
            f'{c + 1}schedule.xlsx')  # открытие файла
        sheet = book.active  # активный лист
        num_cols = sheet.max_column  # количество столбцов
        last_group_cell = 0  # Сколько прошло ячеек от последней группы
        for i in range(6, num_cols):
            if last_group_cell >= 4:  # Если после группы прошло 4 ячейки, ждём следующей группы
                last_group_cell = -1
                continue
            column = []
            for j in range(2, 76):  # Перебираем
                v = str(sheet.cell(column=i, row=j).value)
                if j == 2 and re.match(pattern, v):  # Если ячейка вторая, то проверяем что это номер группы
                    last_group_cell = 0  # Если это так, обнуляем счётчик
                column.append(v)
            if last_group_cell != -1:  # Пока не дошли до следующей группы, не добавляем столбцы,
                schedule_data.append(column)
                last_group_cell += 1


def update_schedule_file():
    """
    Обновляет файл с расписанием

    """
    page = requests.get(MIREA_URL)  # Получаем страницу
    soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    result = soup.find(string="Институт информационных технологий").find_parent("div").find_parent("div").findAll(
        'a', {'class': 'uk-link-toggle'})
    course_pattern = r'([1-3]) курс'
    for i in range(3, 6):
        x = result[i]
        course = x.find('div', 'uk-link-heading').text.lower().strip()
        course_number = re.match(course_pattern, course)
        if course_number:
            course_number = course_number.group(1)
            f = open(f'{course_number}schedule.xlsx', "wb")
            link = x.get('href')
            resp = requests.get(link)
            f.write(resp.content)
            f.close()


if __name__ == '__main__':
    users_groups = []
    schedule_data = []
    vk_session = vk_api.VkApi(
        token='8db4375a8ef6744ac72f08b6145548d26e098e2a7636a4efe397b0976ee4dbd4b202a2f89f5b154239b90')
    vk = vk_session.get_api()
    upload = VkUpload(vk_session)
    db = sqlite3.connect('db.db')
    cur = db.cursor()
    # update_schedule_file()
    schedule_file()
    longpoll = VkLongPoll(vk_session)
    users_to_set_group = set()
    users_to_set_teacher = set()
    users_to_get_teacher = []
    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.text and event.to_me:
            do_command(event.user_id, event.text.lower())
